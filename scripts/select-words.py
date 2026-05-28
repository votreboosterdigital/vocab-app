"""
select-words.py
Extrait les 3000 mots les plus fréquents à l'oral depuis wordFrequency.xlsx (COCA)
Usage: python scripts/select-words.py > scripts/selected_words.json
"""
import json
from openpyxl import load_workbook

XLSX_PATH = 'scripts/wordFrequency.xlsx'
OUTPUT_COUNT = 3000

# PoS retenus : mots apprenables (exclut articles 'a', pronoms 'p', déterminants)
LEARNABLE_POS = {'n', 'v', 'j', 'r', 'i', 'c'}

# Seuils spokPM → CEFR (calibrés sur la distribution réelle du top 3000 oral)
def assign_cefr(spok_pm: float) -> str:
    if spok_pm >= 500: return 'A1'
    if spok_pm >= 150: return 'A2'
    if spok_pm >= 50:  return 'B1'
    if spok_pm >= 20:  return 'B2'
    return 'C1'

wb = load_workbook(XLSX_PATH, read_only=True)
ws = wb['1 lemmas']
rows = list(ws.iter_rows(values_only=True))
# Row 0 = header, data starts at row 1
data = rows[1:]

learnable = []
for r in data:
    pos = r[2]
    word = r[1]
    spok = r[12]
    spok_pm = r[20]
    if pos not in LEARNABLE_POS:
        continue
    if not word or not spok or spok <= 0:
        continue
    word_str = str(word)
    if len(word_str) <= 1:
        continue
    # Exclure les mots avec tirets ou chiffres (health-care, e-mail, mp3...)
    if not word_str.replace('-', '').replace(' ', '').isalpha():
        continue
    learnable.append({
        'word': word_str.lower(),
        'pos': pos,
        'spok': int(spok),
        'spokPM': float(spok_pm) if spok_pm else 0.0,
    })

# Trier par fréquence orale décroissante, prendre les 3000 premiers
learnable_sorted = sorted(learnable, key=lambda x: x['spok'], reverse=True)
top = learnable_sorted[:OUTPUT_COUNT]

# Assigner CEFR et ID
for i, item in enumerate(top, 1):
    item['level'] = assign_cefr(item['spokPM'])
    item['id'] = f"coca-{i:04d}"

# Stats
from collections import Counter
levels = Counter(item['level'] for item in top)
import sys
print(f"✅ Extracted {len(top)} words", file=sys.stderr)
print(f"   A1:{levels['A1']} A2:{levels['A2']} B1:{levels['B1']} B2:{levels['B2']} C1:{levels['C1']}", file=sys.stderr)
print(f"   spokPM range: {top[0]['spokPM']:.0f} → {top[-1]['spokPM']:.1f}", file=sys.stderr)

print(json.dumps(top, indent=2, ensure_ascii=False))
